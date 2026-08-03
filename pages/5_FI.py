# pages/5_FI.py - Focused Improvement Pillar
import os, sys, io
import xlrd
import numpy as np
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
import matplotlib as mpl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ["MPLCONFIGDIR"] = os.path.join(os.getcwd(), ".mplconfig")

from utils.supabase_client import get_supabase

st.set_page_config(page_title="FI Pillar", page_icon="💡", layout="wide")
if "user" not in st.session_state:
    st.switch_page("app.py")

supabase = get_supabase()
role     = st.session_state.get("role", "member")
pillar   = st.session_state.get("pillar", "ALL")
name     = st.session_state.get("name", "User")
can_edit = (role == "plant_manager") or (role == "pillar_leader" and pillar == "FI")

col1, col2, col3 = st.columns([5, 1, 1])
with col1:
    st.markdown("# 💡 Focused Improvement")
    st.markdown(f"Logged in as **{name}** · `{role}` · {'✏️ Full Access' if can_edit else '👁️ View Only'}")
with col3:
    if st.button("🏠 Home", use_container_width=True):
        st.switch_page("pages/1_Home.py")

st.divider()
mpl.rcParams["font.family"] = "DejaVu Sans"
mpl.rcParams["font.size"]   = 11

tab1, tab_setup, tab_projects = st.tabs(["📊 OEE Report & Analysis", "🔁 Setup Analysis", "📋 FI Projects"])

# ════════════════════════════════════════
# SHARED CHART HELPERS
# ════════════════════════════════════════
def _oee_bar_chart(df, oee_col, section_colors, machine_col):
    fig, ax = plt.subplots(figsize=(13.33, 5), dpi=150)
    x = np.arange(len(df))
    colors = [section_colors.get(str(s), "#AAAAAA") for s in df["Section"].fillna("")]
    bars = ax.bar(x, df[oee_col].fillna(0), color=colors, width=0.6, alpha=0.9)
    for bar, val in zip(bars, df[oee_col].fillna(0)):
        ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5,
                f"{val:.1f}%", ha="center", va="bottom", fontsize=9, color="#000000")
    ax.axhline(85, color="#DE201B", linewidth=1.5, linestyle="--", label="Target 85%")
    ax.set_xticks(x); ax.set_xticklabels(df[machine_col].tolist(), rotation=35, ha="right", fontsize=9)
    ax.set_ylabel("OEE (%)"); ax.set_ylim(0, 115)
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False); ax.grid(False)
    from matplotlib.patches import Patch
    patches = [Patch(facecolor=v, label=k) for k,v in section_colors.items() if k in df["Section"].values]
    patches.append(plt.Line2D([0],[0], color="#DE201B", linewidth=1.5, linestyle="--", label="Target 85%"))
    ax.legend(handles=patches, loc="upper center", bbox_to_anchor=(0.5,-0.22), ncol=5, frameon=False, fontsize=9)
    plt.tight_layout(rect=[0,0.10,1,1])
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0); plt.close(fig); return buf

def _arq_chart(df, avl_col, rate_col, qual_col, machine_col):
    fig, ax = plt.subplots(figsize=(13.33, 5), dpi=150)
    x = np.arange(len(df)); w = 0.25
    b1 = ax.bar(x-w, df[avl_col].fillna(0),  w, color="#006394", label="Availability")
    b2 = ax.bar(x,   df[rate_col].fillna(0), w, color="#C1A02E", label="Rate")
    b3 = ax.bar(x+w, df[qual_col].fillna(0), w, color="#D8C37D", label="Quality")
    for b in [b1,b2,b3]:
        for bar in b:
            h = bar.get_height()
            if h > 0:
                ax.text(bar.get_x()+bar.get_width()/2, h+0.5, f"{h:.1f}%",
                        ha="center", va="bottom", fontsize=8, color="#000000")
    ax.set_xticks(x); ax.set_xticklabels(df[machine_col].tolist(), rotation=35, ha="right", fontsize=9)
    ax.set_ylabel("%"); ax.set_ylim(0, 115)
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False); ax.grid(False)
    ax.legend(loc="upper center", bbox_to_anchor=(0.5,-0.22), ncol=3, frameon=False, fontsize=10)
    plt.tight_layout(rect=[0,0.10,1,1])
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0); plt.close(fig); return buf

def _build_excel(df, title, header_color="006394"):
    wb = Workbook(); ws = wb.active; ws.title = title
    hf = PatternFill("solid", fgColor=header_color)
    sf = PatternFill("solid", fgColor="EAF4FB")
    tf = PatternFill("solid", fgColor="FFF9E6")
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    la = Alignment(horizontal="left",   vertical="center")
    th = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"))
    skip = {"Capacity Utilization %","PM & Cleaning (label)","Micro-Stop (label)"}
    headers = [c for c in df.columns if c not in skip]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hf; cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = ca; cell.border = th
    ws.row_dimensions[1].height = 30
    dr = 2; cur_sec = None
    for _, row in df.iterrows():
        sec = str(row.get("Section",""))
        if sec != cur_sec and sec not in ("TOTAL","Unknown",""):
            cur_sec = sec
            cell = ws.cell(row=dr, column=1, value=f"── {sec} ──")
            cell.fill = sf; cell.font = Font(bold=True, size=10); cell.alignment = la
            ws.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=len(headers))
            ws.row_dimensions[dr].height = 18; dr += 1
        is_tot = row.get("Row Type") == "Total"
        for ci, h in enumerate(headers, 1):
            val = row.get(h)
            cell = ws.cell(row=dr, column=ci, value=val)
            cell.border = th; cell.alignment = ca if ci > 2 else la
            cell.fill = tf if is_tot else PatternFill()
            cell.font = Font(bold=True, size=9) if is_tot else Font(size=9)
        ws.row_dimensions[dr].height = 16; dr += 1
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(12, len(str(h))+2)
    ws.column_dimensions["A"].width = 10; ws.column_dimensions["B"].width = 24
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ════════════════════════════════════════
# CORRUGATOR PARSER
# ════════════════════════════════════════
CORR_SECTIONS = {"BHS","FOSBER","SINGLE FACER"}
CORR_M_COLS = [
    "Shift Time","Run Duration (Hrs)","Stop Duration Hrs","Idle Duration Hrs",
    "# of Stops","# of Change Over","# of Quality Change",
    "Linear Meters","Standard Output M","Square Meters",
    "Availability (%)","Rate (%)","Quality (%)","OEE (%)","Avg Max Speed"
]
CORR_T_COLS = [
    "Shift Time","Run Duration (Hrs)","Stop Duration Hrs","Idle Duration Hrs",
    "# of Stops","# of Change Over","# of Quality Change",
    "Linear Meters","Standard Output M","Square Meters",
    "Availability (%)","Rate (%)","Quality (%)","OEE (%)"
]

def parse_corrugator(file_bytes):
    wb = xlrd.open_workbook(file_contents=file_bytes)
    sh = wb.sheet_by_index(0)
    cur_sec = None; rows = []
    for ri in range(sh.nrows):
        row = [sh.cell_value(ri, c) for c in range(sh.ncols)]
        ca  = str(row[0]).strip()
        if ca in CORR_SECTIONS: cur_sec = ca; continue
        if not any(str(v).strip() for v in row): continue
        if isinstance(row[0], str) and ca and ca not in CORR_SECTIONS:
            rec = {"Section": cur_sec, "Machine": ca, "Row Type": "Machine"}
            for i, cn in enumerate(CORR_M_COLS):
                rec[cn] = row[i+1] if (i+1)<len(row) else None
            extras = [(15,"Available Time (Hrs)"),(17,"Average Length/Run"),(19,"Average SQM/Run"),
                      (21,"PM & Cleaning"),(22,"MT"),(23,"Speed Value"),(25,"Average Width/Run"),
                      (27,"Average SQM/Hrs"),(28,"Operator"),(30,"Average Sheet Length"),(32,"Average Sheet Width")]
            for idx, key in extras:
                rec[key] = row[idx] if idx<len(row) else None
            rows.append(rec)
        elif isinstance(row[0], (int, float)) and ca:
            rec = {"Section": cur_sec, "Machine": "TOTAL", "Row Type": "Total"}
            for i, cn in enumerate(CORR_T_COLS):
                rec[cn] = row[i] if i<len(row) else None
            extras = [(15,"Available Time (Hrs)"),(17,"PM & Cleaning"),(19,"Average Length/Run"),
                      (21,"Average SQM/Run"),(22,"MT"),(24,"Average Width/Run"),(26,"Average SQM/Hrs"),
                      (27,"Speed Value"),(29,"# of ProgramID"),(31,"Average Sheet Length"),(33,"Average Sheet Width"),
                      (35,"# of Change Overs"),(37,"Capacity Utilization (%)"),(39,"# of Quality Changes"),
                      (41,"% of Quality Changes"),(43,"Planning Efficiency")]
            for idx, key in extras:
                rec[key] = row[idx] if idx<len(row) else None
            rows.append(rec)
    df = pd.DataFrame(rows)
    skip = {"Section","Machine","Row Type","Operator"}
    for col in df.columns:
        if col not in skip:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

# ════════════════════════════════════════
# CONVERTING PARSER
# ════════════════════════════════════════
CONV_MAP = {
    "BOBST 160-II":"Die Cutters","BOBST 203":"Die Cutters",
    "BOBST MASTERCUT 1":"Die Cutters","BOBST MASTERCUT 2":"Die Cutters",
    "BOBST 924":"FFG","LMC FFG":"FFG","MARTIN 616":"FFG","SATURN":"FFG",
    "IPACK":"Printer","VEGA 2":"Folder Gluers","BAHMÜLLER TURBOX":"Folder Gluers",
    "BAHMULLER STITCHER":"Stitcher","JUMBO":"Jumbo",
    "SINGLE FACER":"Single Facer","SHRINK-WRAPPER":"Shrink Wrapper",
}
CONV_NORM = {k.upper().strip(): v for k,v in CONV_MAP.items()}

def parse_converting(file_bytes, filename=""):
    if filename.lower().endswith(".xlsx"):
        import openpyxl
        wb2 = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws2 = wb2.active
        rows_raw = [[c if c is not None else "" for c in row] for row in ws2.iter_rows(values_only=True)]
    else:
        wb2 = xlrd.open_workbook(file_contents=file_bytes)
        ws2 = wb2.sheet_by_index(0)
        rows_raw = [[ws2.cell_value(r,c) for c in range(ws2.ncols)] for r in range(ws2.nrows)]

    rows = []; tw = None; fqw = None
    for row in rows_raw[2:]:
        ca = str(row[0]).strip() if row else ""
        if not ca: continue
        if "total weight" in ca.lower():
            tw  = row[1] if len(row)>1 else None
            fqw = row[2] if len(row)>2 else None
            break
        if ca.lower().startswith("page"): continue
        sec = CONV_NORM.get(ca.upper(), "Unknown")
        if len([v for v in row[1:] if str(v).strip() not in ("","0","0.0")]) < 3: continue
        col_map = {
            "Hits / Hour":1,"Standard Hits (Rhr)":2,
            "Capacity Utilization %":3,"Capacity Utilization Value":4,
            "Shift Time (Hrs)":5,"Available Time (Hrs)":6,"Run Duration (Hrs)":7,
            "STD Setup (min)":8,"Setup Duration (min)":9,
            "Stop Duration (Hrs)":10,"Idle Duration (Hrs)":11,
            "# of Setup":12,"# of Orders":13,"SQM":14,"Actual Hits":15,
            "Hits / Avl Hour":16,"Hits Rhr+Setup":17,
            "Availability %":18,"Rate %":19,"Quality %":20,"OEE %":21,
            "FG Produced":22,"PM & Cleaning (label)":23,"PM & Cleaning Time":24,
            "Micro-Stop (label)":25,"Micro Stop (min)":26,
            "Unknown Value":27,"Full Qty Weight":28,
        }
        rec = {"Section": sec, "Machine Name": ca, "Row Type": "Machine"}
        for cn, idx in col_map.items():
            rec[cn] = row[idx] if idx<len(row) else None
        rows.append(rec)
    if tw is not None:
        rows.append({"Section":"TOTAL","Machine Name":"TOTAL","Row Type":"Total",
                     "FG Produced": pd.to_numeric(tw, errors="coerce"),
                     "Full Qty Weight": pd.to_numeric(fqw, errors="coerce")})
    df = pd.DataFrame(rows)
    skip = {"Section","Machine Name","Row Type","Capacity Utilization %",
            "PM & Cleaning (label)","Micro-Stop (label)"}
    for col in df.columns:
        if col not in skip:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

# ════════════════════════════════════════
# TAB 1
# ════════════════════════════════════════
# ── FI Projects Tab ──
import importlib.util as _ilu, sys as _sys, os as _os
_fi_proj_path = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), "fi_projects_tab.py")
if _os.path.exists(_fi_proj_path):
    _spec = _ilu.spec_from_file_location("fi_projects_tab", _fi_proj_path)
    _fi_mod = _ilu.module_from_spec(_spec); _spec.loader.exec_module(_fi_mod)
    with tab_projects:
        _fi_mod.render_fi_projects_tab(supabase, role, pillar, name)
else:
    with tab_projects:
        st.warning("fi_projects_tab.py not found in project root.")

with tab1:
    st.markdown("### 📊 OEE Report & Analysis")
    st.caption("Upload one or both OEE report files, then click Generate Analysis.")

    _uc1, _uc2 = st.columns(2)
    oee_file  = _uc1.file_uploader("📂 Corrugator OEE Report (.xls/.xlsx)",
                                    type=["xls","xlsx"], key="fi_oee_file")
    conv_file = _uc2.file_uploader("📂 Converting OEE Report (.xls/.xlsx)",
                                    type=["xls","xlsx"], key="fi_conv_file")
    if not oee_file and not conv_file:
        st.info("⬆️ Upload at least one OEE report file above.")

    if st.button("🔄 Generate OEE Analysis", type="primary", key="fi_oee_run"):
        st.session_state["fi_run_oee"] = True

    if st.session_state.get("fi_run_oee") and (oee_file or conv_file):

        # ── CORRUGATOR ──
        if oee_file:
            st.divider()
            st.markdown("## 🏭 Corrugator OEE")
            try:
                with st.spinner("Parsing Corrugator file..."):
                    df_corr = parse_corrugator(oee_file.getvalue())
                _cm = df_corr[df_corr["Row Type"]=="Machine"]
                st.success(f"✅ Corrugator: {len(_cm)} machines extracted")

                _cc1, _cc2 = st.columns([2,5])
                _scs = _cc1.selectbox("Filter Section",
                    ["All"]+sorted(df_corr["Section"].dropna().unique().tolist()), key="fi_corr_sec")
                _sct = _cc2.multiselect("Row Type", ["Machine","Total"],
                    default=["Machine","Total"], key="fi_corr_type")
                _dfcv = df_corr.copy()
                if _scs != "All": _dfcv = _dfcv[_dfcv["Section"]==_scs]
                if _sct:          _dfcv = _dfcv[_dfcv["Row Type"].isin(_sct)]

                # Summary
                st.markdown("#### OEE Summary")
                _ct = df_corr[df_corr["Row Type"]=="Total"]
                if not _ct.empty:
                    _tcols = st.columns(len(_ct))
                    for i,(_, tr) in enumerate(_ct.iterrows()):
                        _tcols[i].markdown(f"**{tr.get('Section','')}**")
                        _tcols[i].metric("OEE",                  f"{tr.get('OEE (%)',0) or 0:.1f}%")
                        _tcols[i].metric("Availability",          f"{tr.get('Availability (%)',0) or 0:.1f}%")
                        _tcols[i].metric("Rate",                  f"{tr.get('Rate (%)',0) or 0:.1f}%")
                        _tcols[i].metric("Quality",               f"{tr.get('Quality (%)',0) or 0:.1f}%")
                        cap = tr.get("Capacity Utilization (%)", None)
                        _tcols[i].metric("Capacity Utilization",  f"{cap:.1f}%" if pd.notna(cap) else "—")

                st.markdown("#### Extracted Data")
                st.dataframe(_dfcv, use_container_width=True, hide_index=True)

                CORR_COLORS = {"BHS":"#006394","FOSBER":"#C1A02E","SINGLE FACER":"#D8C37D"}
                if not _cm.empty and "OEE (%)" in _cm.columns:
                    st.markdown("#### OEE by Machine")
                    st.image(_oee_bar_chart(_cm, "OEE (%)", CORR_COLORS, "Machine"))
                    st.markdown("#### Availability / Rate / Quality")
                    st.image(_arq_chart(_cm, "Availability (%)","Rate (%)","Quality (%)", "Machine"))

                st.markdown("#### Export")
                st.download_button("📥 Download Corrugator OEE (.xlsx)",
                    data=_build_excel(df_corr,"Corrugator OEE","006394"),
                    file_name=f"Corrugator_OEE_{oee_file.name.replace('.xls','').replace('.xlsx','')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="fi_corr_dl")
            except Exception as e:
                st.error(f"Corrugator Error: {e}")

        # ── CONVERTING ──
        if conv_file:
            st.divider()
            st.markdown("## 🏭 Converting OEE")
            try:
                with st.spinner("Parsing Converting file..."):
                    df_conv = parse_converting(conv_file.getvalue(), conv_file.name)
                _vm = df_conv[df_conv["Row Type"]=="Machine"]
                st.success(f"✅ Converting: {len(_vm)} machines extracted")

                _vs = st.selectbox("Filter Section",
                    ["All"]+sorted([s for s in df_conv["Section"].dropna().unique()
                                    if s not in ("TOTAL","Unknown")]), key="fi_conv_sec")
                _dfvv = _vm.copy()
                if _vs != "All": _dfvv = _dfvv[_dfvv["Section"]==_vs]

                # ── Add computed Capacity Utilization column ──
                if "Available Time (Hrs)" in _vm.columns and "Shift Time (Hrs)" in _vm.columns:
                    _vm = _vm.copy()
                    _vm["Capacity Utilization %"] = (
                        _vm["Available Time (Hrs)"] / _vm["Shift Time (Hrs)"] * 100
                    ).round(1)

                def _wavg(df, val_col, weight_col="Full Qty Weight"):
                    if val_col not in df.columns or weight_col not in df.columns:
                        return None
                    _w = pd.to_numeric(df[weight_col], errors="coerce").fillna(0)
                    _v = pd.to_numeric(df[val_col],    errors="coerce").fillna(0)
                    return (_v * _w).sum() / _w.sum() if _w.sum() > 0 else None

                # Summary
                st.markdown("#### OEE Summary by Section (Weighted by Full Qty Weight)")
                _sec_ord = ["Die Cutters","FFG","Printer","Folder Gluers",
                            "Stitcher","Jumbo","Single Facer","Shrink Wrapper"]
                _active_secs = [s for s in _sec_ord if not _vm[_vm["Section"]==s].empty]
                _scols = st.columns(min(4, len(_active_secs)))
                _si = 0
                for _sec in _sec_ord:
                    _sdf = _vm[_vm["Section"]==_sec]
                    if _sdf.empty: continue
                    _sc = _scols[_si % len(_scols)]
                    _sc.markdown(f"**{_sec}**")
                    for metric, col in [("OEE","OEE %"),("Availability","Availability %"),
                                        ("Rate","Rate %"),("Quality","Quality %"),
                                        ("Cap. Util.","Capacity Utilization %")]:
                        _v = _wavg(_sdf, col)
                        _sc.metric(metric, f"{_v:.1f}%" if _v is not None and pd.notna(_v) else "—")
                    _si += 1

                _tr = df_conv[df_conv["Row Type"]=="Total"]
                if not _tr.empty:
                    _tw1, _tw2 = st.columns(2)
                    _tw1.metric("Total FG Produced", f"{_tr['FG Produced'].iloc[0]:,.0f}")
                    _tw2.metric("Full Qty Weight",   f"{_tr['Full Qty Weight'].iloc[0]:,.0f}")

                st.markdown("#### Extracted Data")
                _disp = [c for c in _dfvv.columns if c not in
                         {"Capacity Utilization %","PM & Cleaning (label)","Micro-Stop (label)"}]
                st.dataframe(_dfvv[_disp], use_container_width=True, hide_index=True)

                CONV_COLORS = {
                    "Die Cutters":"#9B59B6","FFG":"#006394","Printer":"#C1A02E",
                    "Folder Gluers":"#D8C37D","Stitcher":"#E74C3C","Jumbo":"#27AE60",
                    "Single Facer":"#1ABC9C","Shrink Wrapper":"#E67E22","Unknown":"#AAAAAA",
                }
                if not _vm.empty and "OEE %" in _vm.columns:
                    st.markdown("#### OEE by Machine")
                    st.image(_oee_bar_chart(_vm, "OEE %", CONV_COLORS, "Machine Name"))
                    st.markdown("#### Availability / Rate / Quality")
                    st.image(_arq_chart(_vm, "Availability %","Rate %","Quality %", "Machine Name"))

                st.markdown("#### Export")
                st.download_button("📥 Download Converting OEE (.xlsx)",
                    data=_build_excel(df_conv,"Converting OEE","9B59B6"),
                    file_name=f"Converting_OEE_{conv_file.name.replace('.xls','').replace('.xlsx','')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="fi_conv_dl")
            except Exception as e:
                st.error(f"Converting Error: {e}")

        # ══ SPEED BY FLUTE ══
        st.divider()
        st.markdown("## ⚡ Speed by Flute Analysis")
        st.caption("Upload the Speed by Flute .xls file to analyze speed and waste by flute type per machine.")
        flute_file = st.file_uploader("📂 Speed by Flute Report (.xls)", type=["xls"], key="fi_flute_file")

        if flute_file:
            if st.button("🔄 Parse Speed by Flute", type="primary", key="fi_flute_run"):
                st.session_state["fi_run_flute"] = True
                if "fi_flute_df" in st.session_state:
                    del st.session_state["fi_flute_df"]

            if st.session_state.get("fi_run_flute"):
                try:
                    with st.spinner("Parsing Speed by Flute file..."):
                        COL_MAP = {
                            "Gross MT":5,"% Dist MT":9,"Norm Trim":14,"Extra Trim":17,
                            "Up Down":19,"Sheet Waste":23,"Total Waste":25,"Waste %":29,
                            "LM":32,"Gross SQM":37,"Net SQM":40,"Good Boards SQM":42,
                            "Good Boards MT":46,"GSM":50,"Roll Size":55,"Avg Speed":58,
                        }
                        _wb = xlrd.open_workbook(file_contents=flute_file.getvalue())
                        _sh = _wb.sheet_by_index(0)

                        def _safe(ri, ci):
                            try: return _sh.cell_value(ri, ci)
                            except: return None

                        def _has_numeric(ri):
                            for ci in COL_MAP.values():
                                v = _safe(ri, ci)
                                if isinstance(v, (int, float)) and v not in (0, 0.0): return True
                            return False

                        def _row_texts(ri):
                            return [str(_safe(ri,c)).strip() for c in range(min(_sh.ncols,62))
                                    if str(_safe(ri,c)).strip() not in ("","0","0.0","nan","None")]

                        all_rows = []; cur_mach = None; cur_flute = None
                        i = 8
                        while i < _sh.nrows:
                            texts = _row_texts(i)
                            if not texts: i+=1; continue
                            ca = str(_safe(i,0)).strip()
                            cd = str(_safe(i,3)).strip() if _sh.ncols>3 else ""
                            # Machine name row
                            if ca and ca not in ("","nan","None") and not _has_numeric(i):
                                if "%" not in ca and "gross" not in ca.lower():
                                    cur_mach = ca; cur_flute = None; i+=1; continue
                            # Flute name row (col D)
                            if cd and cd not in ("","nan","None") and not _has_numeric(i):
                                if "%" not in cd:
                                    cur_flute = cd; i+=1; continue
                            # Skip % gross rows
                            if any("%" in t and "gross" in t.lower() for t in texts):
                                i+=1; continue
                            # Data / total row
                            if _has_numeric(i) and cur_mach:
                                if not cur_flute:
                                    i+=1; continue  # skip total rows
                                rec = {"Machine": cur_mach, "Flute Type": cur_flute, "Row Type": "Data"}
                                for cn, ci in COL_MAP.items():
                                    v = _safe(i, ci)
                                    rec[cn] = float(v) if isinstance(v,(int,float)) else None
                                cur_flute = None
                                all_rows.append(rec)
                            i+=1

                        df_flute = pd.DataFrame(all_rows)
                        st.session_state["fi_flute_df"] = df_flute
                except Exception as e:
                    st.error(f"Speed by Flute Error: {e}")

            if "fi_flute_df" in st.session_state:
                df_flute = st.session_state["fi_flute_df"]
                _fl_m = df_flute["Machine"].unique().tolist()
                st.success(f"✅ {len(df_flute)} rows across {len(_fl_m)} machines")

                # Debug: show raw % Dist MT values
                with st.expander("🔍 Debug — Raw % Dist MT values from file"):
                    st.dataframe(df_flute[["Machine","Flute Type","% Dist MT","Avg Speed","GSM","Roll Size"]],
                                 use_container_width=True, hide_index=True)

                _sel_m = st.selectbox("Select Machine", ["All"]+_fl_m, key="fi_flute_mach")
                _dffl  = df_flute[df_flute["Machine"]==_sel_m] if _sel_m!="All" else df_flute.copy()
                _data  = _dffl[_dffl["Row Type"]=="Data"]
                _flutes_all = df_flute[df_flute["Row Type"]=="Data"]["Flute Type"].unique().tolist()
                _machs  = df_flute[df_flute["Row Type"]=="Data"]["Machine"].unique().tolist()
                _pal    = ["#006394","#C1A02E","#D8C37D","#9B59B6","#27AE60"]

                st.markdown("#### Extracted Data")
                st.dataframe(_dffl, use_container_width=True, hide_index=True)

                def _flute_bar_chart(metric, ylabel):
                    fig, ax = plt.subplots(figsize=(13.33,5),dpi=150)
                    x = np.arange(len(_flutes_all))
                    w = 0.8/max(len(_machs),1)
                    for mi, mach in enumerate(_machs):
                        mdf = df_flute[(df_flute["Machine"]==mach)&(df_flute["Row Type"]=="Data")]
                        vals = []
                        for f in _flutes_all:
                            sub = mdf[mdf["Flute Type"]==f]
                            vals.append(float(sub[metric].iloc[0]) if not sub.empty and pd.notna(sub[metric].iloc[0]) else 0)
                        offset = (mi - len(_machs)/2 + 0.5)*w
                        bars = ax.bar(x+offset, vals, w, color=_pal[mi%len(_pal)], label=mach, alpha=0.9)
                        for bar, val in zip(bars, vals):
                            if val>0:
                                ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+max(vals)*0.01,
                                        f"{val:.0f}", ha="center", va="bottom", fontsize=8, color="#000000")
                    ax.set_xticks(x); ax.set_xticklabels(_flutes_all, rotation=30, ha="right", fontsize=9)
                    ax.set_ylabel(ylabel)
                    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False); ax.grid(False)
                    ax.legend(loc="upper center", bbox_to_anchor=(0.5,-0.15), ncol=len(_machs), frameon=False, fontsize=9)
                    plt.tight_layout(rect=[0,0.08,1,1])
                    buf = io.BytesIO(); fig.savefig(buf,format="png",dpi=150,bbox_inches="tight")
                    buf.seek(0); plt.close(fig); return buf

                if "Avg Speed" in df_flute.columns:
                    st.divider(); st.markdown("#### Avg Speed by Flute Type")
                    st.image(_flute_bar_chart("Avg Speed","Avg Speed (m/min)"))
                if "Waste %" in df_flute.columns:
                    st.divider(); st.markdown("#### Waste % by Flute Type")
                    st.image(_flute_bar_chart("Waste %","Waste %"))

                st.divider()
                _bfe = io.BytesIO(); df_flute.to_excel(_bfe, index=False, engine="openpyxl"); _bfe.seek(0)
                st.download_button("📥 Download Speed by Flute Table (.xlsx)", data=_bfe,
                    file_name=f"Speed_by_Flute_{flute_file.name.replace('.xls','')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="fi_flute_dl")

                # ══ TARGET SIMULATION ══
                st.divider()
                st.markdown("## 🎯 Production Target Simulation")
                st.caption("Simulate whether the BHS can achieve the MT target given current or adjusted parameters.")

                _df_data = df_flute[df_flute["Row Type"]=="Data"].copy()

                # ── User inputs ──
                _si1, _si2, _si3 = st.columns(3)
                _mt_target = _si1.number_input("Company MT Target", min_value=0.0, value=1000.0,
                                                step=10.0, key="fi_mt_target")
                _waste_pct  = 10.0  # fixed
                _gross_target = _mt_target * (1 + _waste_pct/100)
                _si1.caption(f"Gross MT Target (incl. {_waste_pct}% waste): **{_gross_target:,.1f} MT**")

                _shift_hrs = _si2.number_input("Shift Time (hrs/month)", min_value=1.0,
                                                value=720.0, step=10.0, key="fi_shift_hrs")
                _cu_pct    = _si3.slider("Capacity Utilization %", min_value=50, max_value=100,
                                          value=95, step=1, key="fi_cu_pct")
                _av_pct    = _si3.slider("Availability %", min_value=50, max_value=100,
                                          value=85, step=1, key="fi_av_pct")

                _avail_time = _shift_hrs * (_cu_pct/100) * (_av_pct/100)
                _si2.caption(f"Run Time: **{_avail_time:,.1f} hrs**")

                st.divider()

                # ── Consolidate flutes across machines (BHS + FOSBER → BHS load) ──
                # Sum % distribution by flute type across all machines
                # Sum % Dist MT by flute (already decimal e.g. 0.179 = 17.9%)
                _dist_by_flute = _df_data.groupby("Flute Type")["% Dist MT"].sum().reset_index()
                _dist_by_flute.columns = ["Flute Type","Total % Dist"]
                # % Dist MT is already in percentage form (e.g. 17.9 means 17.9%)

                # For each flute, take weighted avg of GSM, Roll Size, Avg Speed from BHS preferably
                # If flute exists on BHS use BHS values, else use available machine values
                _flute_params = []
                for _, frow in _dist_by_flute.iterrows():
                    ft = frow["Flute Type"]
                    _fdf = _df_data[_df_data["Flute Type"]==ft]
                    _bhs = _fdf[_fdf["Machine"].str.upper().str.contains("BHS")]
                    _src = _bhs if not _bhs.empty else _fdf
                    _gsm       = _src["GSM"].mean()
                    _roll_w    = _src["Roll Size"].mean()
                    _avg_spd   = _src["Avg Speed"].mean()
                    _flute_params.append({
                        "Flute Type":     ft,
                        "Total % Dist":   frow["Total % Dist"],
                        "GSM (avg)":      _gsm,
                        "Roll Width cm (avg)": _roll_w,
                        "Avg Speed m/min (avg)": _avg_spd,
                    })
                _fp_df = pd.DataFrame(_flute_params)

                # ── Baseline speed table ──
                st.markdown("#### 📋 Current Speeds from File (Baseline)")
                _bdisp = _fp_df[["Flute Type","Avg Speed m/min (avg)","GSM (avg)","Roll Width cm (avg)"]].copy()
                _bdisp = _bdisp.rename(columns={
                    "Avg Speed m/min (avg)": "Current Avg Speed (m/min)",
                    "GSM (avg)":             "GSM",
                    "Roll Width cm (avg)":   "Roll Width (cm)",
                })
                for c in ["Current Avg Speed (m/min)","GSM","Roll Width (cm)"]:
                    if c in _bdisp.columns:
                        _bdisp[c] = _bdisp[c].round(1)
                st.dataframe(_bdisp, use_container_width=True, hide_index=True)
                st.divider()

                # ── Speed inputs ──
                st.markdown("#### ⚡ Speed Simulation")
                _glob_col1, _glob_col2 = st.columns([1, 4])
                _speed_increase_pct = _glob_col1.number_input(
                    "Increase ALL speeds by %", min_value=-50, max_value=200,
                    value=0, step=5, key="fi_spd_global_pct"
                )
                _glob_col1.caption("Applies proportionally to all flute speeds from the file.")

                # Compute effective speed per flute = file speed × (1 + global%)
                _speed_overrides = {}
                for _, frow in _fp_df.iterrows():
                    ft = frow["Flute Type"]
                    base_spd = float(frow["Avg Speed m/min (avg)"]) if pd.notna(frow["Avg Speed m/min (avg)"]) else 100.0
                    _speed_overrides[ft] = round(base_spd * (1 + _speed_increase_pct / 100), 1)

                # Show effective speeds as metrics
                _eff_cols = _glob_col2.columns(min(5, len(_speed_overrides)))
                for i, (ft, spd) in enumerate(_speed_overrides.items()):
                    base = float(_fp_df[_fp_df["Flute Type"]==ft]["Avg Speed m/min (avg)"].iloc[0])
                    _eff_cols[i % len(_eff_cols)].metric(
                        ft, f"{spd:.0f} m/min",
                        delta=f"{spd-base:+.0f}" if _speed_increase_pct != 0 else None
                    )

                # ── Calculate per flute ──
                st.divider()
                st.markdown("#### 📊 Simulation Results")

                sim_rows = []
                for _, frow in _fp_df.iterrows():
                    ft       = frow["Flute Type"]
                    dist_pct = frow["Total % Dist"] or 0
                    gsm      = frow["GSM (avg)"]
                    roll_w   = frow["Roll Width cm (avg)"]
                    spd      = _speed_overrides.get(ft, frow["Avg Speed m/min (avg)"] or 100)

                    # Step 1: Gross MT for this flute
                    # % Dist MT is a percentage (e.g. 17.9) — divide by 100
                    gross_mt = _gross_target * (dist_pct / 100) if dist_pct else 0

                    # Step 2: Expected SQM
                    exp_sqm = (gross_mt * 1_000_000) / gsm if gsm and gsm > 0 else 0

                    # Step 3: Expected LM
                    roll_m = (roll_w or 0) * 0.01  # cm → m
                    exp_lm = exp_sqm / roll_m if roll_m > 0 else 0

                    # Step 4: Time needed (hrs)
                    time_needed = (exp_lm / (spd * 60)) if spd and spd > 0 else 0

                    # Net Run Time = LM / (optimal 250 m/min × 60 sec) — time at perfect speed
                    net_run_time_flute = exp_lm / (250 * 60) if exp_lm > 0 else 0
                    _cur_spd = float(frow["Avg Speed m/min (avg)"]) if pd.notna(frow["Avg Speed m/min (avg)"]) and frow["Avg Speed m/min (avg)"] > 0 else 0
                    sim_rows.append({
                        "Flute Type":                  ft,
                        "% Distribution":              round(dist_pct, 2),
                        "Expected Metric Ton":         round(gross_mt, 1),
                        "Expected SQM":                round(exp_sqm, 0),
                        "Expected LM":                 round(exp_lm, 0),
                        "Current Avg Speed (m/min)":   round(_cur_spd, 1),
                        "Proposed Speed (m/min)":      spd,
                        "Time at Current Speed (hrs)": round(exp_lm / (_cur_spd * 60) if _cur_spd > 0 else 0, 2),
                        "Time at Proposed Speed (hrs)":round(time_needed, 2),
                        "Net Run Time (hrs)":          round(net_run_time_flute, 4),
                    })

                sim_df = pd.DataFrame(sim_rows)

                # ── Summary metrics ──
                _total_run_time  = sim_df["Time at Proposed Speed (hrs)"].sum()
                _net_run_time    = sim_df["Net Run Time (hrs)"].sum()
                # Performance Rate = Net Run Time / Total Run Time at proposed speed
                _perf_rate       = (_net_run_time / _total_run_time * 100) if _total_run_time > 0 else 0
                # OEE = Performance Rate × Availability %
                _required_oee    = _perf_rate * (_av_pct / 100)
                # Feasibility = total run time must fit within available run time
                _feasible_all    = _total_run_time <= _avail_time
                _gap             = _avail_time - _total_run_time

                _sm1, _sm2, _sm3, _sm4, _sm5, _sm6 = st.columns(6)
                _sm1.metric("Total Run Time Needed",  f"{_total_run_time:,.1f} hrs")
                _sm2.metric("Net Run Time",            f"{_net_run_time:,.1f} hrs")
                _sm3.metric("Run Time Available",      f"{_avail_time:,.1f} hrs")
                _sm4.metric("Performance Rate",        f"{_perf_rate:.1f}%")
                _sm5.metric("Required OEE",            f"{_required_oee:.1f}%")
                _sm6.metric("Achievable?",             "✅ Yes" if _feasible_all else "❌ No",
                            delta=f"{_gap:+.1f} hrs",
                            delta_color="normal" if _feasible_all else "inverse")

                if not _feasible_all:
                    st.warning(
                        f"⚠️ Total run time needed **{_total_run_time:,.1f} hrs** exceeds "
                        f"available run time **{_avail_time:,.1f} hrs** by **{abs(_gap):,.1f} hrs**. "
                        f"Try increasing speed or availability %."
                    )
                else:
                    st.success(
                        f"✅ Target achievable — **{_gap:,.1f} hrs** of run time remaining. "
                        f"Performance Rate: **{_perf_rate:.1f}%** × Availability: **{_av_pct}%** "
                        f"= Required OEE: **{_required_oee:.1f}%**"
                    )

                # ── Results table ──
                st.dataframe(sim_df, use_container_width=True, hide_index=True)

                # ── Time comparison chart per flute ──
                _fig_sim, _ax_sim = plt.subplots(figsize=(13.33, 5), dpi=150)
                _xs = np.arange(len(sim_df)); _w = 0.35
                _b1 = _ax_sim.bar(_xs-_w/2, sim_df["Time at Current Speed (hrs)"], _w,
                                   color="#D8C37D", label="Current Speed", alpha=0.9)
                _b2 = _ax_sim.bar(_xs+_w/2, sim_df["Time at Proposed Speed (hrs)"], _w,
                                   color="#006394", label="Proposed Speed", alpha=0.9)
                for _b in [_b1, _b2]:
                    for _bar in _b:
                        _h = _bar.get_height()
                        if _h > 0:
                            _ax_sim.text(_bar.get_x()+_bar.get_width()/2, _h+0.1,
                                         f"{_h:.1f}h", ha="center", va="bottom",
                                         fontsize=8, color="#000000")
                _ax_sim.set_xticks(_xs)
                _ax_sim.set_xticklabels(sim_df["Flute Type"].tolist(), rotation=30, ha="right", fontsize=9)
                _ax_sim.set_ylabel("Time to Produce (hrs)")
                _ax_sim.spines["top"].set_visible(False); _ax_sim.spines["right"].set_visible(False)
                _ax_sim.grid(False)
                _ax_sim.legend(loc="upper center", bbox_to_anchor=(0.5,-0.15),
                               ncol=2, frameon=False, fontsize=10)
                plt.tight_layout(rect=[0,0.08,1,1])
                _buf_sim = io.BytesIO()
                _fig_sim.savefig(_buf_sim, format="png", dpi=150, bbox_inches="tight")
                _buf_sim.seek(0); st.image(_buf_sim); plt.close(_fig_sim)

                # ── Download simulation ──
                _buf_simx = io.BytesIO(); sim_df.to_excel(_buf_simx, index=False, engine="openpyxl"); _buf_simx.seek(0)
                st.download_button("📥 Download Simulation Results (.xlsx)", data=_buf_simx,
                    file_name="OEE_Target_Simulation.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="fi_sim_dl")


# ════════════════════════════════════════
# SETUP ANALYSIS TAB
# ════════════════════════════════════════

MACHINE_GROUPS = {
    "FFG": ["LMC", "Martin", "Saturn", "BOBST 924"],
    "Corrugator": ["BHS", "FOSBER"],
    "Die-Cut": ["BOBST 2", "BOBST 3", "MC1", "MC2"],
    "Pre-Print": ["CI4", "CI6"],
    "Folder Gluer": ["VEGA", "TURBOX"],
    "Other": ["JUMBO", "IPAK"],
}
ALL_MACHINES = [m for grp in MACHINE_GROUPS.values() for m in grp]

def _parse_setup_file(file_bytes, filename):
    """Parse a machine order sequence file. Only needs Month and Articleref."""
    if filename.lower().endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes))
    else:
        df = pd.read_excel(io.BytesIO(file_bytes))

    df.columns = [str(c).strip() for c in df.columns]

    # Find Month column (exact match preferred)
    month_col = next((c for c in df.columns if c.strip().lower() == "month"), None)
    if not month_col:
        month_col = next((c for c in df.columns if "month" in c.lower()), None)

    # Find Article column
    article_col = next((c for c in df.columns if c.strip().lower() in
                        ["articleref", "article ref", "article_ref", "ft", "item"]), None)
    if not article_col:
        article_col = next((c for c in df.columns if "article" in c.lower() or "ref" in c.lower()), None)

    # Find Start Date for ordering
    start_col = next((c for c in df.columns if "start" in c.lower()), None)

    if not article_col:
        raise ValueError(f"Cannot find Article/FT column. Columns found: {list(df.columns)}")
    if not month_col:
        raise ValueError(f"Cannot find Month column. Columns found: {list(df.columns)}")

    df["_article"] = df[article_col].astype(str).str.strip()

    # Month column may be stored as datetime in Excel — convert to month name
    raw_month = df[month_col]
    if pd.api.types.is_datetime64_any_dtype(raw_month):
        df["_month"] = raw_month.dt.strftime("%B")
    else:
        # Try converting string to datetime first, then extract month name
        def _to_month_name(val):
            s = str(val).strip()
            # Already a month name
            import calendar
            month_names = [m for m in calendar.month_name if m]
            if s.capitalize() in month_names:
                return s.capitalize()
            # Try parsing as datetime
            try:
                return pd.to_datetime(s).strftime("%B")
            except Exception:
                return s
        df["_month"] = raw_month.apply(_to_month_name)

    df = df[df["_article"].str.len() > 0]
    df = df[df["_month"].str.len() > 0]
    df = df.reset_index(drop=True)
    return df


def _count_setups(df):
    """
    Count setups per article per month.
    Logic: consecutive rows with same article = same run (1 setup).
    Non-consecutive same article in same month = repeated setup.
    Returns: dict {month: {article: n_setups}}
    """
    results = {}  # {month: {article: count}}

    for month, mdf in df.groupby("_month"):
        mdf = mdf.reset_index(drop=True)
        article_setups = {}
        prev_article = None

        for _, row in mdf.iterrows():
            art = row["_article"]
            if art != prev_article:
                # New run for this article
                article_setups[art] = article_setups.get(art, 0) + 1
            prev_article = art

        results[month] = article_setups

    return results


def _build_setup_summary(setup_counts, months):
    """
    Build summary table:
    Rows: #of orders, #of setups, 1, 2, 3, 4, 5, 6+
    Columns: months
    """
    rows = []
    for month in months:
        counts = setup_counts.get(month, {})
        n_orders = len(counts)
        n_setups = sum(counts.values())
        # Distribution: how many FTs had exactly N setups
        dist = {}
        for art, n in counts.items():
            dist[n] = dist.get(n, 0) + 1

        max_setups = max(dist.keys()) if dist else 1
        rows.append({
            "month": month,
            "n_orders": n_orders,
            "n_setups": n_setups,
            **{str(i): dist.get(i, 0) for i in range(1, max_setups + 1)},
        })
    return rows


def _setup_bar_chart(pivot_df, machine_name):
    """Bar chart of repeated setups (2+) per month."""
    months = [c for c in pivot_df.columns if c not in ["Metric"]]
    repeat_rows = [str(i) for i in range(2, 7) if str(i) in pivot_df["Metric"].values]
    if not repeat_rows or not months:
        return None

    colors = ["#006394", "#C1A02E", "#D8C37D", "#9B59B6", "#DE201B"]
    fig, ax = plt.subplots(figsize=(10, 5), dpi=150)
    x = np.arange(len(months))
    w = 0.8 / max(len(repeat_rows), 1)

    for i, row_label in enumerate(repeat_rows):
        row_data = pivot_df[pivot_df["Metric"] == row_label]
        if row_data.empty: continue
        vals = [float(row_data[m].iloc[0]) if m in row_data.columns else 0 for m in months]
        offset = (i - len(repeat_rows) / 2 + 0.5) * w
        bars = ax.bar(x + offset, vals, w, color=colors[i % len(colors)],
                      label=f"{row_label}x setups", alpha=0.9)
        for bar, val in zip(bars, vals):
            if val > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.2,
                        str(int(val)), ha="center", va="bottom", fontsize=9)

    ax.set_xticks(x); ax.set_xticklabels(months, fontsize=10)
    ax.set_ylabel("# of FTs"); ax.set_title(f"{machine_name} — Repeated Setups", fontsize=12)
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.12), ncol=len(repeat_rows), frameon=False)
    plt.tight_layout(rect=[0, 0.08, 1, 1])
    buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0); plt.close(fig)
    return buf


def _detail_table(setup_counts, months):
    """Return per-FT setup counts as a DataFrame."""
    all_arts = sorted(set(art for m in months for art in setup_counts.get(m, {}).keys()))
    rows = []
    for art in all_arts:
        row = {"FT": art}
        total = 0
        for month in months:
            n = setup_counts.get(month, {}).get(art, 0)
            row[month] = n
            total += n
        row["Total"] = total
        rows.append(row)
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("Total", ascending=False).reset_index(drop=True)
    return df


MONTH_ORDER = ["January","February","March","April","May","June",
               "July","August","September","October","November","December"]


with tab_setup:
    st.markdown("### 🔁 Setup Analysis")
    st.caption("Upload order sequence files per machine to identify repeated setups (same FT set up multiple times in a month).")

    # ── File uploaders per group ──────────────────────────────────────────────
    uploaded_files = {}  # {machine_name: uploaded_file}

    for group_name, machines in MACHINE_GROUPS.items():
        with st.expander(f"📂 {group_name}", expanded=True):
            cols = st.columns(len(machines))
            for i, machine in enumerate(machines):
                f = cols[i].file_uploader(
                    machine,
                    type=["xls", "xlsx", "csv"],
                    key=f"setup_file_{machine}",
                    label_visibility="visible"
                )
                if f:
                    uploaded_files[machine] = f

    n_uploaded = len(uploaded_files)
    if n_uploaded == 0:
        st.info("⬆️ Upload at least one machine file above, then click Generate.")
    else:
        st.success(f"✅ {n_uploaded} file(s) uploaded: {', '.join(uploaded_files.keys())}")

    if st.button("🔄 Generate Setup Analysis", type="primary", key="setup_btn", disabled=n_uploaded == 0):
        st.session_state["setup_run"] = True
        st.session_state["setup_results"] = {}

    if st.session_state.get("setup_run") and uploaded_files:
        results = {}  # {machine: {month: {article: n}}}
        errors  = {}

        with st.spinner("Parsing files..."):
            for machine, f in uploaded_files.items():
                try:
                    df = _parse_setup_file(f.getvalue(), f.name)
                    results[machine] = _count_setups(df)
                except Exception as e:
                    errors[machine] = str(e)

        if errors:
            for m, e in errors.items():
                st.error(f"❌ {m}: {e}")

        if not results:
            st.warning("No data could be parsed. Check your file formats.")
        else:
            # Collect all months across all machines, sorted
            all_months_set = set()
            for machine_counts in results.values():
                all_months_set.update(machine_counts.keys())
            months = sorted(all_months_set, key=lambda m: MONTH_ORDER.index(m) if m in MONTH_ORDER else 99)

            # ── Summary table (all machines) ──────────────────────────────────
            st.divider()
            st.markdown("### 📋 Summary — All Machines")

            summary_rows = []
            for group_name, machines in MACHINE_GROUPS.items():
                for machine in machines:
                    if machine not in results: continue
                    setup_counts = results[machine]
                    machine_rows = _build_setup_summary(setup_counts, months)
                    for mr in machine_rows:
                        summary_rows.append({"Group": group_name, "Machine": machine, **mr})

            if summary_rows:
                # Build pivot display matching the screenshot format
                display_rows = []
                prev_machine = None
                for sr in summary_rows:
                    machine_label = sr["Machine"] if sr["Machine"] != prev_machine else ""
                    prev_machine  = sr["Machine"]

                display_data = {}
                for machine in [m for grp in MACHINE_GROUPS.values() for m in grp if m in results]:
                    machine_rows = [sr for sr in summary_rows if sr["Machine"] == machine]
                    if not machine_rows: continue

                    display_data[machine] = {}
                    for mr in machine_rows:
                        m = mr["month"]
                        display_data[machine].setdefault(m, {})
                        display_data[machine][m]["#of orders"] = mr["n_orders"]
                        display_data[machine][m]["#of setups"] = mr["n_setups"]
                        for i in range(1, 7):
                            display_data[machine][m][str(i)] = mr[str(i)]

                # Build the pivot DataFrame
                metric_labels = ["#of orders", "#of setups", "1", "2", "3", "4", "5", "6"]
                pivot_rows = []
                for machine, mdata in display_data.items():
                    for label in metric_labels:
                        row = {"Machine": machine, "Metric": label}
                        for month in months:
                            row[month] = mdata.get(month, {}).get(label, 0)
                        pivot_rows.append(row)

                pivot_df = pd.DataFrame(pivot_rows)

                # Style and display
                def _style_pivot(df):
                    styled = df.style
                    def _row_style(row):
                        if row["Metric"] == "#of orders":
                            return ["background-color: #EAF3FB; font-weight: bold"] * len(row)
                        elif row["Metric"] == "#of setups":
                            return ["background-color: #FFF3E0; font-weight: bold"] * len(row)
                        elif row["Metric"] in ["3","4","5","6"]:
                            return ["color: #DE201B; font-weight: bold"] + [""] * (len(row)-1)
                        return [""] * len(row)
                    styled = styled.apply(_row_style, axis=1)
                    return styled

                st.dataframe(_style_pivot(pivot_df), use_container_width=True, hide_index=True)

                # ── Excel export of summary ───────────────────────────────────
                _buf_sum = io.BytesIO()
                pivot_df.to_excel(_buf_sum, index=False, engine="openpyxl")
                _buf_sum.seek(0)
                st.download_button(
                    "📥 Download Summary (.xlsx)",
                    data=_buf_sum,
                    file_name="Setup_Analysis_Summary.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="setup_summary_dl"
                )

            # ── Per-machine detail ────────────────────────────────────────────
            st.divider()
            st.markdown("### 🔍 Detail by Machine")

            for group_name, machines in MACHINE_GROUPS.items():
                group_machines = [m for m in machines if m in results]
                if not group_machines: continue

                st.markdown(f"#### {group_name}")
                for machine in group_machines:
                    setup_counts = results[machine]
                    machine_rows = _build_setup_summary(setup_counts, months)
                    if not machine_rows: continue

                    with st.expander(f"🔧 {machine}", expanded=True):
                        # Mini summary metrics
                        metric_cols = st.columns(len(months))
                        for i, month in enumerate(months):
                            mr = next((r for r in machine_rows if r["month"] == month), None)
                            if mr:
                                metric_cols[i].metric(f"{month[:3]} Orders", mr["n_orders"])
                                metric_cols[i].metric(f"{month[:3]} Setups", mr["n_setups"])
                                repeat_count = sum(mr[str(k)] for k in range(2, 7))
                                metric_cols[i].metric(f"{month[:3]} Repeated", repeat_count,
                                                      delta=f"{repeat_count/mr['n_orders']*100:.0f}% of orders" if mr["n_orders"] else None,
                                                      delta_color="inverse" if repeat_count > 0 else "off")

                        # Chart
                        # Find max setup count across all months for this machine
                        _max_s = max(
                            (int(k) for r in machine_rows for k in r.keys()
                             if k.isdigit() and r[k] > 0),
                            default=6
                        )
                        machine_pivot = pd.DataFrame([
                            {"Metric": label,
                             **{month: next((r.get(label, 0) for r in machine_rows if r["month"] == month), 0)
                                for month in months}}
                            for label in [str(i) for i in range(1, _max_s + 1)]
                        ])
                        chart_buf = _setup_bar_chart(machine_pivot, machine)
                        if chart_buf:
                            st.image(chart_buf)

                        # Detail table (all FTs)
                        st.markdown("##### FT-level Detail")
                        detail_df = _detail_table(setup_counts, months)
                        if not detail_df.empty:
                            # Highlight FTs with repeated setups
                            def _highlight_repeats(row):
                                month_vals = [row.get(m, 0) for m in months]
                                if any(v >= 3 for v in month_vals):
                                    return ["background-color: #FFEAEA"] * len(row)
                                elif any(v >= 2 for v in month_vals):
                                    return ["background-color: #FFF9E0"] * len(row)
                                return [""] * len(row)
                            st.dataframe(
                                detail_df.style.apply(_highlight_repeats, axis=1),
                                use_container_width=True,
                                hide_index=True,
                                height=min(400, 35 * len(detail_df) + 38)
                            )

                            # Download per machine
                            _buf_m = io.BytesIO()
                            detail_df.to_excel(_buf_m, index=False, engine="openpyxl")
                            _buf_m.seek(0)
                            st.download_button(
                                f"📥 Download {machine} FT Detail (.xlsx)",
                                data=_buf_m,
                                file_name=f"Setup_Analysis_{machine.replace(' ','_')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"setup_dl_{machine}"
                            )

                # ── Cross-machine combined table (only if 2+ machines uploaded in group) ──
                if len(group_machines) >= 2:
                    st.markdown(f"##### 🔀 {group_name} — Combined FT Analysis (All Machines)")
                    st.caption("FTs summed across all uploaded machines in this group. Highlights FTs that ran on multiple machines in the same month.")

                    combined = {}
                    machine_presence = {}

                    for machine in group_machines:
                        setup_counts = results[machine]
                        for month, art_counts in setup_counts.items():
                            for ft, n in art_counts.items():
                                if ft not in combined:
                                    combined[ft] = {}
                                    machine_presence[ft] = {}
                                combined[ft][month] = combined[ft].get(month, 0) + n
                                if month not in machine_presence[ft]:
                                    machine_presence[ft][month] = []
                                machine_presence[ft][month].append(machine)

                    all_fts = sorted(combined.keys())
                    comb_rows = []
                    for ft in all_fts:
                        row = {"FT": ft}
                        total = 0
                        on_multiple = False
                        for month in months:
                            n = combined[ft].get(month, 0)
                            row[month] = n
                            total += n
                            if len(machine_presence[ft].get(month, [])) > 1:
                                on_multiple = True
                        row["Total"] = total
                        row["On Multiple Machines"] = "⚠️ Yes" if on_multiple else ""
                        comb_rows.append(row)

                    comb_df = pd.DataFrame(comb_rows)

                    # Build summary header rows: #of orders and #of setups
                    summary_header = []
                    for metric, label in [("n_orders", "#of orders"), ("n_setups", "#of setups")]:
                        hrow = {"FT": label, "Total": "", "On Multiple Machines": ""}
                        for month in months:
                            if metric == "n_orders":
                                # Count FTs that had at least 1 setup this month
                                hrow[month] = sum(1 for ft in all_fts if combined[ft].get(month, 0) > 0)
                            else:
                                # Sum all setups this month
                                hrow[month] = sum(combined[ft].get(month, 0) for ft in all_fts)
                        summary_header.append(hrow)

                    if not comb_df.empty:
                        header_df = pd.DataFrame(summary_header)
                        comb_df = pd.concat([header_df, comb_df], ignore_index=True)
                    if not comb_df.empty:
                        comb_df = comb_df.sort_values("Total", ascending=False).reset_index(drop=True)

                        def _highlight_combined(row):
                            if row.get("FT") == "#of orders":
                                return ["background-color: #EAF3FB; font-weight: bold"] * len(row)
                            if row.get("FT") == "#of setups":
                                return ["background-color: #FFF3E0; font-weight: bold"] * len(row)
                            if row.get("On Multiple Machines") == "⚠️ Yes":
                                return ["background-color: #FFF3CD"] * len(row)
                            if isinstance(row.get("Total"), (int, float)) and row.get("Total", 0) >= 3:
                                return ["background-color: #FFEAEA"] * len(row)
                            return [""] * len(row)

                        st.dataframe(
                            comb_df.style.apply(_highlight_combined, axis=1),
                            use_container_width=True,
                            hide_index=True,
                            height=min(500, 35 * len(comb_df) + 38)
                        )

                        _buf_comb = io.BytesIO()
                        comb_df.to_excel(_buf_comb, index=False, engine="openpyxl")
                        _buf_comb.seek(0)
                        st.download_button(
                            f"📥 Download {group_name} Combined FT Analysis (.xlsx)",
                            data=_buf_comb,
                            file_name=f"Setup_Analysis_{group_name.replace(' ','_')}_Combined.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"setup_dl_combined_{group_name}"
                        )